import streamlit as st
import time
# import camelot.io as camelot
import camelot
import PyPDF2
import os
# import base64
from datetime import datetime
from streamlit import session_state
from st_aggrid import AgGrid, GridUpdateMode, JsCode
from st_aggrid.grid_options_builder import GridOptionsBuilder
# import re
import math
import glob
import pandas as pd
from openpyxl import load_workbook
import shutil
from request import (
        get_symbols,
        add_company,
        add_file,
        get_financial_words_col,
        get_financial_words_row,
        get_json_financial_format,
        get_json_format,
        insert_data,
        get_allFiles
    )
from extraction.pdf_to_image import (convert_file)
from extraction.aws_image import (image_extraction)
from extraction.saving import (save_json_to_db)
from streamlit_extras.switch_page_button import switch_page

# Initialization
if 'pg_input' not in session_state:
    session_state['pg_input'] = ''

if 'status' not in session_state:
    session_state['status'] = False

if 'og_uploaded_file' not in session_state:
    session_state['og_uploaded_file'] = None

currency = ""
fiscal_month = ""
duplicate_num_format = []
confirm_headers_list = []
# confirm_rows_list = []
confirm_search_col_list = []
search_col_list_check = []
financial_format =[]
number_format = []
dataframe_list = []
is_df_empty_list = []
button_clicked = False

if 'dataframes' not in session_state:
    session_state['dataframes'] = []

if 'column_input' not in st.session_state:
    st.session_state['column_input'] = False

if 'column_del' not in st.session_state:
    st.session_state['column_del'] = False

if "upload_file_status" not in st.session_state:
    session_state['upload_file_status'] = True

# re-initalise again to update previous false
session_state['upload_file_status'] = True

if ("com_name" and "selected_comName" and "com_id" and "selected_comID") not in st.session_state:
    st.session_state['com_name'] = ""
    st.session_state['com_id'] = ""
    st.session_state['selected_comName'] = ""
    session_state['selected_comID'] = ""
    session_state['upload_file_status'] = False
    st.error("Please upload a file for extraction.", icon="🚨")

if "text_option" not in st.session_state:
    session_state['text_option'] = False

# to persist the form when user clicks on submit after previewing data
if "extract_state" not in session_state:
    session_state["extract_state"] = False

# retrieve from upload files page
com_name = session_state["com_name"]
com_id = session_state["com_id"]
selected_comName = session_state["selected_comName"]
selected_comID = session_state["selected_comID"]

temp_path = "./temp_files"
dir = os.listdir(temp_path)

# check if file was uploaded
if st.session_state['text_option'] == True and session_state['upload_file_status'] and len(dir) > 1:
    st.header(com_name)
else:
    st.header(selected_comName)

number = [            
            "Unable to Determine",
            "Thousand",
            "Million",
            "Billion",
            "Trillion",
            "Quadrillion"
        ]

month = [
            "Not Selected",
            "January",
            "February",
            "March",
            "April",
            "May",
            "June",
            "July",
            "August",
            "September",
            "October",
            "November",
            "December"
        ]

def get_file_name (file):
    filename = os.path.splitext(file)[0]
    return filename

def get_file_type (file):
    filetype = os.path.splitext(file)[1]
    return filetype

def get_total_pgs_PDF (file):
    file = open(file, 'rb')
    pdf = PyPDF2.PdfReader(file)
    pages = len(pdf.pages)
    return pages

def check_tables_single_PDF (file):
    tables = camelot.read_pdf(file, pages="1", flavor="stream", edge_tol=100, row_tol=10)
    return tables

def check_tables_multi_PDF (file, pages):
    tables = camelot.read_pdf(file, pages=pages, flavor="stream", edge_tol=100, row_tol=10)
    return (tables)

def get_number_format (excel_path):
    num_array = []
    workbook = load_workbook(excel_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"Title = {sheet.title}")
        for value in sheet.iter_rows(values_only=True):
            for cell in value:
                if cell is None:
                    continue

                if "thousand" in str(cell).lower():
                    num_array.append(1)

                if "million" in str(cell).lower():
                    num_array.append(2)

                if "billion" in str(cell).lower():
                    num_array.append(3)

                if "trillion" in str(cell).lower():
                    num_array.append(4)

                if "quadrillion" in str(cell).lower():
                    num_array.append(5)            
    
    index = 0
    if len(num_array) > 0:
        # get the first index because it is the first to retrieve
        index = num_array[0]
    else:
        index = 0 # unable to determine
    
    return index

def sort_num_list(index):
    format = number.pop(index)
    number.insert(0, format)
    return number

def get_currency_list():
    currency_acronyms = []
    symbols = get_symbols()
    for key in symbols:
        currency_acronyms.append(key + " (" + symbols[key] + ")")
    currency_acronyms.insert(0, "Not Selected")        
    return currency_acronyms

def image_viewer(dataframes):

    # check if dataframe is empty
    if len(dataframes) < 1:
        st.error('Please upload an image with a table.', icon="🚨")

    else:
        extraction_container = st.empty()
        with extraction_container.container():
            for i in range(len(dataframes)):
                # if dataframe is not empty (manage to extract some things out)        
                statement, format, is_df_empty, search_col_check, confirm_headers, search_col = viewer_func(dataframes[i], i, 'img', "", "") 


def viewer_func(df, num, id, num_form, convert):
    search_col_check = []

    # to loop through and search for number format
    df.to_excel("./temp_files/" + str(num) + ".xlsx")    

    # for displaying in aggrid 
    df.to_csv("./temp_files/" + str(num) + ".csv")
    dataframe = pd.read_csv("./temp_files/" + str(num) + ".csv")

    option = ""
    selected = ""

    st.subheader('Extracted Table ' + str(num+1))
    delete = st.checkbox("Don't extract Table " + str(num+1) + "? (Clicking this will refresh the entire page and changes made might be lost.)")

    # check if an empty dataframe is extracted
    if dataframe.empty:
        is_df_empty = True
        is_df_empty_list.append(True)
        df_empty_msg = "Empty dataframe, please upload a pdf with a filled table or Try AWS."
        st.error(df_empty_msg, icon="🚨")
    else:
        is_df_empty = False
        is_df_empty_list.append(False)
        
        col1, col2 = st.columns(2)

        # financial statement ddl
        with col1:
            if not delete:
                option = st.selectbox('Select a Financial Statement:', ('Not Selected', 'Income Statement', 'Balance Sheet', 'Cash Flow'), key=id+str(num))
                financial_format.append(option)
                if option == "Not Selected":
                    st.warning("Financial Statement is a required field", icon="⭐")
            else:
                option = st.selectbox('Select a Financial Statement:', ('Not Selected', 'Income Statement', 'Balance Sheet', 'Cash Flow'), key=id+str(num), disabled=True)
        
        # number format ddl
        with col2:
            num_format = get_number_format("./temp_files/" + str(num) + ".xlsx")
            new_num_list = sort_num_list(num_format)
            options = list(range(len(new_num_list)))
            numFormat = ""
            if not delete:
                if num_form!="":
                    st.selectbox("Number Format:", [num_form], key="num_format -" + id + str(num), disabled=True)
                    numFormat = num_form
                    st.info("You cannot choose a different number format. Number Format should be consistent for every table in each uploaded report.", icon="ℹ️")

                elif num_form =="" and convert=="pdfimg":
                    
                    nums= [            
                            "Unable to Determine",
                            "Thousand",
                            "Million",
                            "Billion",
                            "Trillion",
                            "Quadrillion"
                        ]
                    extracted_num_format = nums[num_format]


                    del nums[num_format]
                    nums.insert(0, extracted_num_format)
                    numFormat = st.selectbox("Number Format:", nums, key="num_format -" + id + str(num))
                    if numFormat == "Unable to Determine":
                        st.warning("Number Format is a required field.", icon="⭐")
                else:
                    i = st.selectbox("Number Format:", options, format_func=lambda x: new_num_list[int(x)], key="format -" + id + str(num))
                    number_format.append(new_num_list[i])
                    numFormat = new_num_list[i]
                    if new_num_list[i] == "Unable to Determine":
                        st.warning("Number Format is a required field.", icon="⭐")
                    
                    # each financial statement should have a consistent number format
                    different = 0
                    for saved_num in number_format:
                        if new_num_list[i] != saved_num:
                            different += 1

                    if different > 0:
                        st.warning("You cannot choose a different number format. Number Format should be consistent for each table in each uploaded report.", icon="⭐")
                        duplicate_num_format.append(True)
            else:
                i = st.selectbox("Number Format:", options, format_func=lambda x: new_num_list[int(x)], key="format -" + id + str(num), disabled=True)



        st.subheader("Edit Headers")

        column1, column2 = st.columns(2)

        with column1:
            if not delete:
                options = st.multiselect('Select Columns to Delete:', list(dataframe.columns), key="coldelete -" + id + str(num))
                st.session_state['column_del'] = True
                    
                for col_option in options:
                    dataframe.drop(col_option, axis=1, inplace=True)
            else:
                options = st.multiselect('Select Columns to Delete:', list(dataframe.columns), key="coldelete -" + id + str(num), disabled=True)

                
        with column2:
            renamecol_tooltip = "Column headers must be unique and strictly years or years+quarters. For yearly statements, you can rename columns which falls under a specified year as 2020_1, 2020_2 etc. For quarterly statements, you can rename the headers as 2020 Q1, 2020 Q2 etc. If there are other columns which falls under 2020 Q1 for instance, you can rename it as 2020 Q1_1"
            if not delete:
                options = st.multiselect('Select Column Header(s) to Rename:', list(dataframe.columns), help=renamecol_tooltip, key="colrename -" + id + str(num))
                
                for i in range(len(options)):
                    old_name = options[i]
                    column_name = st.text_input("Enter New Header Name for "+ str(options[i]), value=options[i], key="table -" + str(num) + str(i))
                    st.session_state['column_input'] = True
                    
                    if column_name in (dataframe.columns) and old_name !=column_name:
                        st.error("Header name already exisit. Try a different name.")
                    else:
                        dataframe.rename(columns = {options[i]: column_name}, inplace = True) 
            else:
                options = st.multiselect('Select Column Header(s) to Rename:', list(dataframe.columns), help=renamecol_tooltip, key="colrename -" + id + str(num), disabled=True)

        # get the columns to search financial keywords
        column_headers = list(dataframe.columns)
        confirm_headers_tooltip = "Select the columns with all rows consisting of financial keywords in your word dictionary e.g. Revenue, Liabilities, Operating Net Cash Flow etc."
        if not delete:
            confirm_headers = st.multiselect('Select the Column(s) with Financial Statement Keywords:', column_headers, column_headers[0], help=confirm_headers_tooltip ,key="confirm_headers -" + id + str(num))
        else:
            confirm_headers = st.multiselect('Select the Column(s) with Financial Statement Keywords:', column_headers, column_headers[0], help=confirm_headers_tooltip ,key="confirm_headers -" + id + str(num), disabled=True)
        

        confirm_headers_list.append(confirm_headers)
        
        if len(confirm_headers_list[num]) < 1:
            if not delete:
                st.error("You need to select at least 1 column header.", icon="🚨")

        # display aggrid
        if not delete:
            delete_row = JsCode("""
                function(e) {
                    let api = e.api;
                    let sel = api.getSelectedRows();
                    api.applyTransaction({remove: sel})    
                };
                """)  

            string_to_add_row = "\n\n function(e) { \n \
            let api = e.api; \n \
            let rowIndex = e.rowIndex + 1; \n \
            api.applyTransaction({addIndex: rowIndex, add: [{}]}); \n \
                }; \n \n"
            
            cell_button_add = JsCode('''
                class BtnAddCellRenderer {
                    init(params) {
                        this.params = params;
                        this.eGui = document.createElement('div');
                        this.eGui.innerHTML = `
                        <span>
                            <style>
                            .btn_add {
                            border: none;
                            color: black;
                            text-align: center;
                            text-decoration: none;
                            display: inline-block;
                            font-size: 10px;
                            font-weight: bold;
                            height: 2.5em;
                            width: 8em;
                            cursor: pointer;
                            }

                            .btn_add :hover {
                            background-color: #05d588;
                            }
                            </style>
                            <button id='click-button' 
                                class="btn_add" 
                                >Add Row Below</button>
                        </span>
                    `;
                    }

                    getGui() {
                        return this.eGui;
                    }

                };
                ''')
            
            st.info('To Delete Row(s): Click the checkbox', icon="ℹ️")

            gd = GridOptionsBuilder.from_dataframe(dataframe)
            gd.configure_default_column(editable=True,groupable=True)
            gd.configure_column("", onCellClicked=JsCode(string_to_add_row), cellRenderer=cell_button_add, editable=False, autoHeight=True,lockPosition='left')
            gd.configure_selection(selection_mode= 'multiple', use_checkbox=True)
            gd.configure_grid_options(onRowSelected = delete_row, pre_selected_rows=[])

            gridOptions = gd.build()
            grid_table = AgGrid(dataframe, 
                        gridOptions = gridOptions, 
                        enable_enterprise_modules = True,
                        fit_columns_on_grid_load = True,
                        # update_mode= GridUpdateMode.MANUAL,
                        update_mode = GridUpdateMode.VALUE_CHANGED | GridUpdateMode.SELECTION_CHANGED,
                        editable = True,
                        height= 450,
                        allow_unsafe_jscode=True)    

            # retrieve edited table & append
            new_df = grid_table['data']
            dataframe_list.append(new_df)

        else:
            AgGrid(dataframe,
                   enable_enterprise_modules = True,
                   fit_columns_on_grid_load = True,
                   editable = False,
                   height= 450) 

        # get columns headers of where cell is located
        search_col_list = list(dataframe.columns)
        for item in confirm_headers:
            search_col_list.remove(item)
        
        search_col = []
        if not delete:
            search_headers = st.multiselect('Select the Column(s) to Search Through:', search_col_list, key="search_cols -" + id + str(num))
            confirm_search_col_list.append(search_headers)
            search_col.append(search_headers)

            if len(search_headers) <= 0:
                st.warning("You need to select at least 1 column to locate cell value.", icon="⭐")
                search_col_list_check.append(False)
                search_col_check.append(False)
            else:
                search_col_list_check.append(True)
                search_col_check.append(True)
        
        else:
            search_headers = st.multiselect('Select the Column(s) to Search Through:', search_col_list, key="search_cols -" + id + str(num), disabled=True)
    
    print("=====HERE" + str(num)+ "========")
    print(number_format)
    return (option, numFormat, is_df_empty, search_col_check, confirm_headers, search_col)

def extract_tables (tables):
    # CHECK ACCURACY
    accuracy = []
    financial_format = []
    number_format = []
    confirm_headers_list = []
    # confirm_rows_list = []
    confirm_search_col_list = []
    search_col_list_check = []
    if len(tables) == 0:
        st.error('Please upload a pdf with a table.', icon="🚨")
    else:
        for i in range(len(tables)):
            accuracy.append(tables[i].parsing_report["accuracy"])
        if (any(i < 75 for i in accuracy)):
            print(accuracy)
            dfs = convert_file()
            for i in range(len(dfs[0])):
                statement, format, is_df_empty, search_col_list_check, confirm_headers, search_col = viewer_func(dfs[i], i, "pdfimg", "", "")
        else:
            for i in range(len(tables)):
                statement, format, is_df_empty, search_col_list_check, confirm_headers, search_col = viewer_func(tables[i], i, 'camelot', "", "")

# make sure a file was being uploaded first
if session_state['upload_file_status'] == True:
    # if temp_files is not empty then extract
    if len(dir) > 1:

        pg_input = session_state.pg_input
        status = session_state.status
        num = 0 
        file_paths = glob.glob("./temp_files/*")
        count = 0
        totalpages = 0
        dataframes = []
        is_image = False
        path_list=[]

       
        for path in (file_paths):
            # print("PATH: "  + path)
            # print("PATH: "  + str(path == "./temp_files\selected_pages.pdf"))
            # if path!="./temp_files\selected_pages.pdf":
                file_type = get_file_type(path)

                # file is pdf
                if file_type == '.pdf':               
                    file_path = glob.glob("./temp_files/*.pdf")[0]
                    session_state["uploaded_file"] = file_path
                    file_name = get_file_name(file_path)
                    totalpages = get_total_pgs_PDF(file_path)
                    num+=1

                #file is image
                elif file_type == '.png' or file_type == '.jpg' or file_type == '.jpeg' and file_type != '.txt':
                    file_path = glob.glob("./temp_files/*" + file_type)
                    file_name = get_file_name(file_path[0])
                    # is_image.append(True)
                    is_image = True
                    if file_path not in path_list:
                        dataframes = image_extraction(file_path[0])
                        for file in file_path:
                            # dataframes.append(image_extraction(file))
                            path_list.append(file_path)
                            print(image_extraction(file))  

        # at least 1 page
        if (totalpages > 0 or is_image == True):
            # Check file type
            try_aws_btn = st.button("Try AWS (Switch Page)")
            if try_aws_btn and file_path.endswith(".pdf"):
                switch_page("try aws") 
            
            st.subheader('Basic Form Data')
            col1, col2 = st.columns(2)
            
            with col1:
                currency_list = get_currency_list()
                option = st.selectbox('Select a Currency:', currency_list, key="currency_singlepg_pdf")
                if option != "Not Selected":
                    currency = option
                else:
                    st.warning("Currency is a required field", icon="⭐")
            
            # fiscal month ddl
            with col2:
                month_list = list(range(len(month)))
                selected = st.selectbox("Fiscal Start Month:", month_list, format_func=lambda x: month[int(x)], key="fiscalmnth")
                if selected != 0:
                    fiscal_month = selected
                else:
                    st.warning("Fiscal Start Month is a required field.", icon="⭐")

            if (is_image == True):
                image_viewer(dataframes)

            # single page pdf
            if (is_image == False):
                if (totalpages == 1):

                    # # try aws button
                    # button_clicked = False
                    # btn_placeholder = st.empty()
                    # with btn_placeholder.container():
                    #     # if session_state["status"]:
                    #         if (st.button("Try AWS", key="aws_singlepg_pdf")):
                    #             origin = './temp_files/'
                    #             target = './selected_files/'
                    #             files = os.listdir(origin)
                    #             files_target = os.listdir(target)
                    #             for file in files_target:
                    #                 if file=="file.pdf":
                    #                     os.remove(target+file)
                    #             for file in files:
                    #                 if file!="test.txt" and file.endswith(".pdf"):
                    #                     file_type = get_file_type(file)
                    #                     shutil.copy(origin+file, target)
                    #                     os.rename(target+file, target+"file"+file_type)

                    #             button_clicked = True
                    #             btn_placeholder.empty()
                        
                    tables = check_tables_single_PDF(file_path)
                    extraction_container = st.empty()
                    with extraction_container.container():
                        extract_tables(tables)
                
                # multi page pdf
                else:
                    # user input is successful on page 3
                    if (status == True and pg_input != ''):
                        # try aws button 
                        # button_clicked = False
                        # btn_placeholder = st.empty()
                        # with btn_placeholder.container():
                        #     if session_state["status"]:
                        #         if (st.button("Try AWS", key="aws_multipg_pdf"+str(num))):
                        #             button_clicked = True
                        #             btn_placeholder.empty()

                        tables = check_tables_multi_PDF(file_path, str(pg_input))
                        extraction_container = st.empty()
                        with extraction_container.container():
                            extract_tables(tables)
                        
                    else:
                        if (session_state['upload_file_status'] == True):
                            st.error("Please specify the pages you want to extract.", icon="🚨")
            
            # if button_clicked:
            #     extraction_container.empty()
            #     next_extraction = st.empty()
            #     with next_extraction.container():
            #         dfs = convert_file()
            #         for i in range(len(dfs)):
            #             # fixed from dfs[0][i] for multi page pdf
            #             statement, format, is_df_empty = viewer_func(dfs[i][0], i, "btnclicked")
        
        # if at least 1 dataframe is not empty
        if False in is_df_empty_list:
            search_col_list_check
            currency
            fiscal_month
            financial_format
            number_format
            duplicate_num_format
            confirm_headers_list
            confirm_search_col_list
            # show extract button
            if st.button("Extract", key="extract") or session_state["extract_state"]:
                # save extract button session
                session_state["extract_state"] = True

                # saving function here
                save_json_to_db(dataframe_list, search_col_list_check, currency, fiscal_month, financial_format, number_format, duplicate_num_format, confirm_headers_list, confirm_search_col_list)

        else:
            st.error("Nothing was extracted from all the tables. Please try again later or Try AWS.", icon="🚨")                        

    # no files was uploaded
    else:
        st.error('Please upload a file for extraction.', icon="🚨")